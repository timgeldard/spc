"""
SPC Export router.

Endpoints:
  POST /api/spc/export  — generate Excel or CSV download for scorecard / chart data / signals

Returns StreamingResponse with appropriate Content-Type and Content-Disposition.
PDF export is handled client-side via window.print().
"""

import io
import json
from typing import Optional

from fastapi import APIRouter, Header, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, TypeAdapter, field_validator, model_validator
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from backend.utils.db import (
    check_warehouse_config,
    resolve_token,
)
from backend.utils.rate_limit import limiter
from backend.dal.spc_analysis_dal import fetch_scorecard
from backend.dal.spc_charts_dal import (
    fetch_chart_data,
    fetch_count_chart_data,
    fetch_p_chart_data,
)
from backend.routers.spc_common import handle_sql_error
from backend.schemas.spc_schemas import _validate_date

router = APIRouter()

_MATERIAL_ID_MAX_LEN = 40
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def sanitize_spreadsheet_value(value):
    if isinstance(value, str) and value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def _sanitize_row(row):
    return [sanitize_spreadsheet_value(cell) for cell in row]


class SignalExportEntry(BaseModel):
    rule: Optional[str] = None
    chart: Optional[str] = "X"
    indices: list[int] = []
    description: Optional[str] = None


class ExportRequest(BaseModel):
    export_type: str          # 'excel' | 'csv'
    export_scope: str         # 'scorecard' | 'chart_data' | 'attribute_chart' | 'signals'
    material_id: str
    mic_id: Optional[str] = None
    mic_name: Optional[str] = None
    plant_id: Optional[str] = None
    operation_id: Optional[str] = None
    chart_type: Optional[str] = None
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    signals_json: Optional[str] = None   # JSON-encoded signals list from frontend

    @field_validator("material_id")
    @classmethod
    def check_material_id(cls, v: str) -> str:
        if len(v) > _MATERIAL_ID_MAX_LEN:
            raise ValueError(f"material_id must be at most {_MATERIAL_ID_MAX_LEN} characters")
        return v

    @field_validator("date_from", "date_to")
    @classmethod
    def check_date(cls, v: Optional[str]) -> Optional[str]:
        return _validate_date(v, "date")

    @field_validator("export_type")
    @classmethod
    def check_export_type(cls, v: str) -> str:
        if v not in ("excel", "csv"):
            raise ValueError("export_type must be 'excel' or 'csv'")
        return v

    @field_validator("export_scope")
    @classmethod
    def check_export_scope(cls, v: str) -> str:
        if v not in ("scorecard", "chart_data", "attribute_chart", "signals"):
            raise ValueError("export_scope must be 'scorecard', 'chart_data', 'attribute_chart', or 'signals'")
        return v

    @model_validator(mode="after")
    def validate_signals_json(self) -> "ExportRequest":
        if self.export_scope != "signals" or self.signals_json is None:
            return self
        try:
            raw = json.loads(self.signals_json)
        except (json.JSONDecodeError, TypeError) as exc:
            raise ValueError("signals_json must be valid JSON") from exc
        try:
            TypeAdapter(list[SignalExportEntry]).validate_python(raw)
        except Exception as exc:
            raise ValueError("signals_json must be a list of signal objects") from exc
        return self


# ---------------------------------------------------------------------------
# Excel styling helpers
# ---------------------------------------------------------------------------

HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="1B3A4B")
CENTER_ALIGN = Alignment(horizontal="center")

STATUS_FILLS = {
    "excellent": PatternFill("solid", fgColor="D1FAE5"),
    "good":      PatternFill("solid", fgColor="ECFDF5"),
    "marginal":  PatternFill("solid", fgColor="FFFBEB"),
    "poor":      PatternFill("solid", fgColor="FEF2F2"),
    "grey":      PatternFill("solid", fgColor="F9FAFB"),
}


def _style_header_row(ws, row_num: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = CENTER_ALIGN


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


# ---------------------------------------------------------------------------
# Data fetchers (reuse DAL logic so export and API stay aligned)
# ---------------------------------------------------------------------------

async def _fetch_scorecard(token: str, body: ExportRequest) -> list[dict]:
    return await fetch_scorecard(token, body.material_id, body.plant_id, body.date_from, body.date_to)


async def _fetch_chart_data(token: str, body: ExportRequest) -> list[dict]:
    if not body.mic_id:
        return []
    return await fetch_chart_data(
        token,
        body.material_id,
        body.mic_id,
        body.mic_name,
        body.plant_id,
        body.date_from,
        body.date_to,
        None,
        operation_id=body.operation_id,
    )


async def _fetch_attribute_chart_data(token: str, body: ExportRequest) -> list[dict]:
    if not body.mic_id:
        return []
    if body.chart_type == "p_chart":
        return await fetch_p_chart_data(
            token,
            body.material_id,
            body.mic_id,
            body.mic_name,
            body.plant_id,
            body.date_from,
            body.date_to,
            operation_id=body.operation_id,
        )
    chart_subtype = "u" if body.chart_type == "u_chart" else "np" if body.chart_type == "np_chart" else "c"
    return await fetch_count_chart_data(
        token,
        body.material_id,
        body.mic_id,
        body.mic_name,
        body.plant_id,
        body.date_from,
        body.date_to,
        chart_subtype,
        operation_id=body.operation_id,
    )


# ---------------------------------------------------------------------------
# Excel builders
# ---------------------------------------------------------------------------

def _build_scorecard_excel(rows: list[dict], material_id: str) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Capability Scorecard"

    headers = _scorecard_export_headers()
    ws.append(_sanitize_row(headers))
    _style_header_row(ws, 1, len(headers))

    for row in _scorecard_export_rows(rows):
        ws.append(_sanitize_row(row))
        # Colour status cell (column 15 = "Status")
        status = str(row[-1] or "").strip().lower()
        fill   = STATUS_FILLS.get(status)
        if fill:
            ws.cell(row=ws.max_row, column=15).fill = fill

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_chart_data_excel(rows: list[dict], material_id: str, mic_name: Optional[str]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Measurement Data"

    headers = ["Batch ID", "Batch Date", "Batch Seq", "Sample Seq", "Value", "Nominal", "Tolerance", "Valuation"]
    ws.append(_sanitize_row(headers))
    _style_header_row(ws, 1, len(headers))

    for row in rows:
        ws.append(_sanitize_row([
            row.get("batch_id"),
            row.get("batch_date"),
            row.get("batch_seq"),
            row.get("sample_seq"),
            row.get("value"),
            row.get("nominal"),
            row.get("tolerance"),
            row.get("valuation"),
        ]))

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _attribute_export_headers(chart_type: Optional[str]) -> list[str]:
    if chart_type == "p_chart":
        return ["Batch ID", "Batch Date", "Batch Seq", "Inspected", "Nonconforming", "P Value"]
    if chart_type == "u_chart":
        return ["Batch ID", "Batch Date", "Batch Seq", "Opportunities", "Defects"]
    if chart_type == "np_chart":
        return ["Batch ID", "Batch Date", "Batch Seq", "Inspected", "Defect Count"]
    return ["Batch ID", "Batch Date", "Batch Seq", "Inspected", "Defect Count"]


def _attribute_export_rows(rows: list[dict], chart_type: Optional[str]) -> list[list[object]]:
    export_rows: list[list[object]] = []
    for row in rows:
        if chart_type == "p_chart":
            export_rows.append([
                row.get("batch_id"),
                row.get("batch_date"),
                row.get("batch_seq"),
                row.get("n_inspected"),
                row.get("n_nonconforming"),
                row.get("p_value"),
            ])
            continue
        export_rows.append([
            row.get("batch_id"),
            row.get("batch_date"),
            row.get("batch_seq"),
            row.get("n_inspected"),
            row.get("defect_count"),
        ])
    return export_rows


def _build_attribute_chart_excel(rows: list[dict], chart_type: Optional[str]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attribute Chart Data"

    headers = _attribute_export_headers(chart_type)
    ws.append(_sanitize_row(headers))
    _style_header_row(ws, 1, len(headers))

    for row in _attribute_export_rows(rows, chart_type):
        ws.append(_sanitize_row(row))

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _scorecard_export_headers() -> list[str]:
    return [
        "MIC ID", "Characteristic", "Batches", "Samples", "Mean", "Std Dev",
        "Min", "Max", "Target", "Pp", "Ppk", "Z Score", "DPMO", "OOC Rate", "Status",
    ]


def _scorecard_export_rows(rows: list[dict]) -> list[list[object]]:
    export_rows: list[list[object]] = []
    for row in rows:
        ooc_pct = f"{row.get('ooc_rate', 0) * 100:.1f}%" if row.get("ooc_rate") is not None else "—"
        export_rows.append([
            row.get("mic_id"),
            row.get("mic_name"),
            row.get("batch_count"),
            row.get("sample_count"),
            row.get("mean_value"),
            row.get("stddev_overall"),
            row.get("min_value"),
            row.get("max_value"),
            row.get("nominal_target"),
            row.get("pp"),
            row.get("ppk"),
            row.get("z_score"),
            row.get("dpmo"),
            ooc_pct,
            row.get("capability_status", "").capitalize(),
        ])
    return export_rows


def _parse_signals(signals_json: Optional[str]) -> list[SignalExportEntry]:
    if not signals_json:
        return []
    raw = json.loads(signals_json)
    return TypeAdapter(list[SignalExportEntry]).validate_python(raw)


def _build_signals_excel(signals: list[dict]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Signals Log"

    headers = ["Rule", "Chart", "Point Indices", "Description"]
    ws.append(_sanitize_row(headers))
    _style_header_row(ws, 1, len(headers))

    for sig in signals:
        ws.append(_sanitize_row([
            sig.get("rule"),
            sig.get("chart", "X"),
            ", ".join(str(i) for i in sig.get("indices", [])),
            sig.get("description"),
        ]))

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------

def _rows_to_csv(headers: list[str], rows: list[list]) -> str:
    import csv, io as _io
    buf = _io.StringIO()
    w   = csv.writer(buf)
    w.writerow(_sanitize_row(headers))
    for row in rows:
        w.writerow(_sanitize_row(row))
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Endpoint
# ---------------------------------------------------------------------------

@router.post("/export")
@limiter.limit("20/minute")
async def spc_export(
    request: Request,
    body: ExportRequest,
    x_forwarded_access_token: Optional[str] = Header(default=None),
    authorization: Optional[str] = Header(default=None),
):
    """Generate an Excel or CSV download for SPC data."""
    token = resolve_token(x_forwarded_access_token, authorization)
    check_warehouse_config()

    scope = body.export_scope
    fmt   = body.export_type

    # --- Signals (no DB call needed — passed from frontend) ---
    if scope == "signals":
        signals = _parse_signals(body.signals_json)

        if fmt == "excel":
            buf  = _build_signals_excel([signal.model_dump() for signal in signals])
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=spc_signals.xlsx"},
            )
        else:
            headers_csv = ["Rule", "Chart", "Point Indices", "Description"]
            rows_csv    = [
                [s.rule, s.chart or "X",
                 ", ".join(str(i) for i in s.indices), s.description]
                for s in signals
            ]
            csv_text = _rows_to_csv(headers_csv, rows_csv)
            return StreamingResponse(
                iter([csv_text]),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=spc_signals.csv"},
            )

    # --- Scorecard ---
    if scope == "scorecard":
        try:
            rows = await _fetch_scorecard(token, body)
        except Exception as exc:
            handle_sql_error(exc)

        if fmt == "excel":
            buf = _build_scorecard_excel(rows, body.material_id)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=spc_scorecard.xlsx"},
            )
        else:
            csv_text = _rows_to_csv(_scorecard_export_headers(), _scorecard_export_rows(rows))
            return StreamingResponse(
                iter([csv_text]),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=spc_scorecard.csv"},
            )

    # --- Chart data ---
    if scope == "chart_data":
        try:
            rows = await _fetch_chart_data(token, body)
        except Exception as exc:
            handle_sql_error(exc)

        if fmt == "excel":
            buf = _build_chart_data_excel(rows, body.material_id, body.mic_name)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=spc_chart_data.xlsx"},
            )
        else:
            headers_csv = ["Batch ID", "Batch Date", "Batch Seq", "Sample Seq",
                           "Value", "Nominal", "Tolerance", "Valuation"]
            rows_csv = [
                [r.get("batch_id"), r.get("batch_date"), r.get("batch_seq"),
                 r.get("sample_seq"), r.get("value"), r.get("nominal"),
                 r.get("tolerance"), r.get("valuation")]
                for r in rows
            ]
            csv_text = _rows_to_csv(headers_csv, rows_csv)
            return StreamingResponse(
                iter([csv_text]),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=spc_chart_data.csv"},
            )

    # --- Attribute chart data ---
    if scope == "attribute_chart":
        try:
            rows = await _fetch_attribute_chart_data(token, body)
        except Exception as exc:
            handle_sql_error(exc)

        if fmt == "excel":
            buf = _build_attribute_chart_excel(rows, body.chart_type)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=spc_attribute_chart_data.xlsx"},
            )

        headers_csv = _attribute_export_headers(body.chart_type)
        rows_csv = _attribute_export_rows(rows, body.chart_type)
        csv_text = _rows_to_csv(headers_csv, rows_csv)
        return StreamingResponse(
            iter([csv_text]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=spc_attribute_chart_data.csv"},
        )
